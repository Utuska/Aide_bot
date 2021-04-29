import pandas as pd
from openpyxl import load_workbook

# df = pd.DataFrame({'Name': ['Manchester City', 'Real Madrid', 'Liverpool',
#                             'FC Bayern München', 'FC Barcelona', 'Juventus'],
#                    'League': ['English Premier League (1)', 'Spain Primera Division (1)',
#                               'English Premier League (1)', 'German 1. Bundesliga (1)',
#                               'Spain Primera Division (1)', 'Italian Serie A (1)'],
#                    'TransferBudget': [176000000, 188500000, 90000000,
#                                       100000000, 180500000, 105000000]})

# print(df)


# for item in df:
#     print(item)

# df.to_excel('../documents/Unbe.xlsx')


# fn = r"../documents/user.xlsx"
#
# wb = load_workbook(fn)
# print(wb)
# ws = wb["Sheet1"]
# print(ws)
#
# row = (101, 102, 103)   # <--- новая строка
# ws.append(row)
#
# wb.save(fn)
# wb.close()

def formation_excel():
    records = [(1, 2, 3, 4, 5), (1, 2, 3, 4, 5)]
    date = []
    status = []
    for item in records:
        status.append(item[3])
        date.append(item[2])

    import pandas as pd
    from openpyxl import load_workbook

    df = pd.DataFrame({'Date': date,
                       'Status': status
                       })

    from translate import Translator
    name_user = 'Гаврилы'
    filename = 'Ndfo'
    # translator = Translator(from_lang="ru", to_lang="en")
    # filename = translat or.translate(name_user)

    filepath = f'../documents/{filename}.xlsx'
    df.to_excel(f'../documents/{filename}.xlsx')


# formation_excel()
